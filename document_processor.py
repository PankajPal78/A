import os
import logging
from typing import List, Dict, Any
import PyPDF2
from docx import Document as DocxDocument
import openpyxl
import pandas as pd
import tiktoken
from sentence_transformers import SentenceTransformer
import chromadb
from chromadb.config import Settings
import uuid

logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Handles document processing, chunking, and embedding generation"""
    
    def __init__(self, chroma_db_path: str = "./data/chroma_db"):
        self.chroma_db_path = chroma_db_path
        self.embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
        self.encoding = tiktoken.get_encoding("cl100k_base")
        
        # Initialize ChromaDB
        self.chroma_client = chromadb.PersistentClient(
            path=chroma_db_path,
            settings=Settings(anonymized_telemetry=False)
        )
        
        # Get or create collection
        try:
            self.collection = self.chroma_client.get_collection("document_chunks")
        except:
            self.collection = self.chroma_client.create_collection(
                name="document_chunks",
                metadata={"hnsw:space": "cosine"}
            )
    
    def process_document(self, file_path: str, document_id: int) -> Dict[str, Any]:
        """Process a document and return processing results"""
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            
            # Extract text based on file type
            if file_extension == '.pdf':
                text_content = self._extract_pdf_text(file_path)
            elif file_extension in ['.docx', '.doc']:
                text_content = self._extract_docx_text(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                text_content = self._extract_excel_text(file_path)
            elif file_extension == '.txt':
                text_content = self._extract_txt_text(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_extension}")
            
            # Chunk the text
            chunks = self._chunk_text(text_content)
            
            # Generate embeddings and store in vector DB
            chunk_ids = []
            for i, chunk in enumerate(chunks):
                chunk_id = f"doc_{document_id}_chunk_{i}"
                chunk_ids.append(chunk_id)
                
                # Generate embedding
                embedding = self.embedding_model.encode(chunk['content']).tolist()
                
                # Store in ChromaDB
                self.collection.add(
                    ids=[chunk_id],
                    embeddings=[embedding],
                    documents=[chunk['content']],
                    metadatas=[{
                        'document_id': document_id,
                        'chunk_index': i,
                        'page_number': chunk.get('page_number'),
                        'chunk_size': len(chunk['content'])
                    }]
                )
            
            return {
                'success': True,
                'chunk_count': len(chunks),
                'chunk_ids': chunk_ids,
                'total_chars': sum(len(chunk['content']) for chunk in chunks)
            }
            
        except Exception as e:
            logger.error(f"Error processing document {file_path}: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text_content = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                if page_text.strip():
                    text_content.append({
                        'content': page_text,
                        'page_number': page_num + 1
                    })
        return text_content
    
    def _extract_docx_text(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        doc = DocxDocument(file_path)
        text_content = []
        current_page = 1
        
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_content.append({
                    'content': paragraph.text,
                    'page_number': current_page
                })
                # Simple page break detection (not perfect but works for most cases)
                if '\f' in paragraph.text or 'page break' in paragraph.text.lower():
                    current_page += 1
        
        return text_content
    
    def _extract_excel_text(self, file_path: str) -> str:
        """Extract text from Excel file"""
        text_content = []
        workbook = openpyxl.load_workbook(file_path)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = []
            
            for row in sheet.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    sheet_text.append(row_text)
            
            if sheet_text:
                text_content.append({
                    'content': '\n'.join(sheet_text),
                    'page_number': len(text_content) + 1
                })
        
        return text_content
    
    def _extract_txt_text(self, file_path: str) -> str:
        """Extract text from TXT file"""
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        return [{
            'content': content,
            'page_number': 1
        }]
    
    def _chunk_text(self, text_content: List[Dict[str, Any]], chunk_size: int = 1000, overlap: int = 200) -> List[Dict[str, Any]]:
        """Chunk text into smaller pieces for better retrieval"""
        chunks = []
        
        for page_data in text_content:
            content = page_data['content']
            page_number = page_data.get('page_number', 1)
            
            # Tokenize the content
            tokens = self.encoding.encode(content)
            
            # Create chunks with overlap
            for i in range(0, len(tokens), chunk_size - overlap):
                chunk_tokens = tokens[i:i + chunk_size]
                chunk_text = self.encoding.decode(chunk_tokens)
                
                if chunk_text.strip():
                    chunks.append({
                        'content': chunk_text.strip(),
                        'page_number': page_number
                    })
        
        return chunks
    
    def search_similar_chunks(self, query: str, n_results: int = 5, document_ids: List[int] = None) -> List[Dict[str, Any]]:
        """Search for similar chunks using vector similarity"""
        try:
            # Generate query embedding
            query_embedding = self.embedding_model.encode(query).tolist()
            
            # Prepare where clause for filtering by document IDs
            where_clause = {}
            if document_ids:
                where_clause = {"document_id": {"$in": document_ids}}
            
            # Search in ChromaDB
            results = self.collection.query(
                query_embeddings=[query_embedding],
                n_results=n_results,
                where=where_clause
            )
            
            # Format results
            similar_chunks = []
            for i, (chunk_id, distance, document, metadata) in enumerate(zip(
                results['ids'][0],
                results['distances'][0],
                results['documents'][0],
                results['metadatas'][0]
            )):
                similar_chunks.append({
                    'chunk_id': chunk_id,
                    'content': document,
                    'similarity_score': 1 - distance,  # Convert distance to similarity
                    'metadata': metadata
                })
            
            return similar_chunks
            
        except Exception as e:
            logger.error(f"Error searching similar chunks: {str(e)}")
            return []
    
    def delete_document_chunks(self, document_id: int):
        """Delete all chunks for a specific document"""
        try:
            # Get all chunks for this document
            results = self.collection.get(
                where={"document_id": document_id}
            )
            
            if results['ids']:
                self.collection.delete(ids=results['ids'])
                logger.info(f"Deleted {len(results['ids'])} chunks for document {document_id}")
            
        except Exception as e:
            logger.error(f"Error deleting chunks for document {document_id}: {str(e)}")